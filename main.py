# OTP Relay Server - main.py
# Stack: FastAPI + Python 3.12
# No external APIs. Runs entirely on your company LAN.
#
# Delivery model: OTP is displayed on-screen via polling.

import asyncio
import hashlib
import json
import logging
import os
import re
import secrets
import threading
import urllib.parse
import urllib.request
from collections import deque
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from io import BytesIO
from typing import Any, Dict, List, Optional

import bcrypt
import openpyxl
from dotenv import load_dotenv
from prometheus_client import Counter, Gauge, Histogram, generate_latest, CONTENT_TYPE_LATEST

try:
    import redis
except ImportError:
    redis = None
from fastapi import FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR / "frontend"


def _resolve_runtime_path(value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else BASE_DIR / path


load_dotenv(BASE_DIR / ".env")

app = FastAPI(title="OTP Relay")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # safe for current LAN-only deployment
    allow_methods=["*"],
    allow_headers=["*"],
)

# -- Prometheus metrics -------------------------------------------------------
# These metrics describe the OTP-relay subsystem (the queue, the iPhone path).
# Naming follows docs/dev/naming-conventions.md: otp_* prefix for OTP-relay
# machinery; non-OTP portal features would use rta_* or portal_* later.
#
# Counters are incremented inside audit() via the OTP_AUDIT_COUNTERS table.
# Gauges (queue depth, active user) are measured on demand at scrape time.

otp_claims_total = Counter(
    "otp_claims_total",
    "Total number of times a user claimed the active slot.",
)
otp_delivered_total = Counter(
    "otp_delivered_total",
    "Total number of OTPs successfully delivered to a recipient.",
)
otp_claim_expired_total = Counter(
    "otp_claim_expired_total",
    "Total number of claims that timed out without an OTP arriving.",
)
otp_request_duration_seconds = Histogram(
    "otp_request_duration_seconds",
    "HTTP request duration in seconds.",
    ["method", "endpoint", "status"],
)

# Audit-event-name to counter mapping. Hook is in audit() below.
OTP_AUDIT_COUNTERS = {
    "claim_queued": otp_claims_total,
    "otp_delivered": otp_delivered_total,
    "claim_expired": otp_claim_expired_total,
}

# Gauges measured on demand. The set_function callbacks are invoked by
# prometheus_client at scrape time, so the returned values are always live.
# Defined further down once the underlying queue state is available; see
# _attach_gauge_callbacks() at module-load tail.
otp_queue_depth = Gauge(
    "otp_queue_depth",
    "Current number of OTPs pending delivery (claim queue depth).",
)
otp_active_user = Gauge(
    "otp_active_user",
    "Whether someone currently holds the active slot (1) or not (0).",
)


@app.middleware("http")
async def _record_request_duration(request: Request, call_next):
    """Record per-request latency in the otp_request_duration_seconds histogram.

    The 'endpoint' label is the route's path template (e.g. /claim-status/{token}),
    not the actual URL, so we don't blow up cardinality with one series per
    distinct token. Falls back to the raw path if no route matched (404, etc.).
    """
    import time as _time
    start = _time.perf_counter()
    response = await call_next(request)
    elapsed = _time.perf_counter() - start

    route = request.scope.get("route")
    endpoint = getattr(route, "path", None) or "unknown"
    otp_request_duration_seconds.labels(
        method=request.method,
        endpoint=endpoint,
        status=str(response.status_code),
    ).observe(elapsed)
    return response


# Expose /metrics as a regular route rather than a sub-app mount.
# An app.mount() requires a trailing slash from clients (Prometheus and curl
# alike), which is fiddly. A plain GET route handles /metrics cleanly.
@app.get("/metrics", include_in_schema=False)
def metrics() -> Response:
    return Response(generate_latest(), media_type=CONTENT_TYPE_LATEST)

# -- Config -------------------------------------------------------------------
SMS_SECRET_TOKEN = os.getenv("SMS_SECRET_TOKEN", "changeme")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

# How long the active user has to trigger their OTP before being evicted.
# Other users wait until this window expires or OTP is delivered.
CLAIM_EXPIRY_SEC = int(os.getenv("CLAIM_EXPIRY_SEC", "90"))

# How long the delivered OTP stays visible on-screen before being purged.
OTP_DISPLAY_SEC = int(os.getenv("OTP_DISPLAY_SEC", "285"))  # 4 min 45 sec

# If two claims arrive within this window, log a concurrent_risk event.
CONCURRENT_RISK_SEC = int(os.getenv("CONCURRENT_RISK_SEC", "30"))

USERS_EXCEL_PATH = str(_resolve_runtime_path(os.getenv("USERS_EXCEL_PATH", "data/users.xlsx")))
USERS_EXCEL_MAX_BYTES = int(os.getenv("USERS_EXCEL_MAX_BYTES", str(5 * 1024 * 1024)))
AUDIT_LOG_PATH = str(_resolve_runtime_path(os.getenv("AUDIT_LOG_PATH", "data/audit.log")))
REDIS_URL = os.getenv("REDIS_URL", "").strip()
REDIS_REQUIRED = os.getenv("REDIS_REQUIRED", "0").strip() == "1"

# -- State --------------------------------------------------------------------
users: Dict[str, Dict[str, str]] = {}
claim_queue: deque = deque()

USER_IMPORT_HEADER_ALIASES = {
    "token": "token",
    "user token": "token",
    "username": "token",
    "name": "name",
    "display name": "name",
    "email": "email",
    "email address": "email",
    "mail": "email",
    "test_env": "test_env",
    "test env": "test_env",
    "test environment": "test_env",
    "test_environment": "test_env",
    "testenv": "test_env",
    "prod_env": "prod_env",
    "prod env": "prod_env",
    "prod environment": "prod_env",
    "prod_environment": "prod_env",
    "production env": "prod_env",
    "production environment": "prod_env",
    "prodenv": "prod_env",
}


def _normalize_user_import_header(value: Any) -> str:
    clean = re.sub(r"\s+", " ", str(value or "").strip().lower().replace("-", " ").replace("_", " "))
    return USER_IMPORT_HEADER_ALIASES.get(clean, clean.replace(" ", "_"))


def _xlsx_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()

# Delivered OTPs held in memory only - never written to disk or logs.
# Structure: { token: { "otp": str, "arrived_at": datetime } }
pending_otps: Dict[str, Dict[str, Any]] = {}

# Redis shared-state keys. Redis is used when REDIS_URL is configured and reachable.
# The in-memory structures above remain as a safe fallback while REDIS_REQUIRED=0.
REDIS_QUEUE_KEY = "otp:queue"
REDIS_QUEUE_LOCK_KEY = "otp:lock:queue"
REDIS_CLAIM_PREFIX = "otp:claim:"
REDIS_PENDING_PREFIX = "otp:pending:"
REDIS_ADMIN_SESSION_PREFIX = "admin:session:"
REDIS_ADMIN_LOGIN_ATTEMPT_PREFIX = "admin:login_attempt:"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%SZ",
)
logger = logging.getLogger("otp-relay")
redis_client = None

# -- Server-backed wizard/admin state -----------------------------------------
DATA_DIR = _resolve_runtime_path(os.environ.get("OTP_RELAY_DATA_DIR", "data"))
WIZARD_FILE = DATA_DIR / "wizard_progress.json"
AUTH_FILE = DATA_DIR / "admin_auth.json"
CONFIG_FILE = DATA_DIR / "admin_config.json"
DEFAULT_ADMIN_TOKENS = ["JPR", "AMD", "SCH"]
ADMIN_TTL_SECONDS = 1 * 60 * 60  # 1-hour sliding session for admin users
ADMIN_SESSIONS: Dict[str, Dict[str, Any]] = {}  # session → {"ts": float, "who": str}
ADMIN_LOGIN_ATTEMPTS: Dict[str, Dict[str, Any]] = {}
ADMIN_LOGIN_WINDOW_SECONDS = int(os.getenv("ADMIN_LOGIN_WINDOW_SECONDS", "300"))
ADMIN_LOGIN_MAX_ATTEMPTS = int(os.getenv("ADMIN_LOGIN_MAX_ATTEMPTS", "8"))
ADMIN_LOGIN_LOCKOUT_SECONDS = int(os.getenv("ADMIN_LOGIN_LOCKOUT_SECONDS", "900"))
# PIN reset codes — stored in memory (+ Redis when available) with a short TTL
ADMIN_RESET_CODES: Dict[str, Dict[str, Any]] = {}  # token → {"code": str, "expires": float}
REDIS_ADMIN_RESET_PREFIX = "admin:reset:"
ADMIN_RESET_TTL_SECONDS = 15 * 60  # codes expire after 15 minutes
WIZARD_DB_LOCK = threading.Lock()
WIZARD_CLIENT_SECRET_MIN_LENGTH = int(os.getenv("WIZARD_CLIENT_SECRET_MIN_LENGTH", "32"))


def _utcnow_naive() -> datetime:
    """Return a UTC timestamp compatible with existing naive queue timestamps."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _redis_enabled() -> bool:
    return bool(REDIS_URL)


def _init_redis_client() -> None:
    global redis_client

    if not REDIS_URL:
        redis_client = None
        logger.info("Redis disabled; using in-memory runtime state")
        return

    if redis is None:
        redis_client = None
        message = "REDIS_URL is set but redis package is not installed"
        if REDIS_REQUIRED:
            raise RuntimeError(message)
        logger.warning(message)
        return

    redis_client = redis.Redis.from_url(
        REDIS_URL,
        decode_responses=True,
        socket_connect_timeout=2,
        socket_timeout=2,
        health_check_interval=30,
    )

    try:
        redis_client.ping()
        logger.info("Redis connected: %s", REDIS_URL)
    except Exception as exc:
        redis_client = None
        if REDIS_REQUIRED:
            raise RuntimeError(f"Redis is required but not reachable: {exc}") from exc
        logger.warning("Redis not reachable; continuing with in-memory runtime state: %s", exc)


def _redis_status() -> str:
    if not REDIS_URL:
        return "disabled"

    if redis_client is None:
        return "unavailable"

    try:
        redis_client.ping()
        return "ok"
    except Exception:
        return "error"


def _use_redis_state() -> bool:
    return redis_client is not None


def _redis_claim_key(token: str) -> str:
    return f"{REDIS_CLAIM_PREFIX}{token}"


def _redis_pending_key(token: str) -> str:
    return f"{REDIS_PENDING_PREFIX}{token}"


def _redis_admin_session_key(session: str) -> str:
    return f"{REDIS_ADMIN_SESSION_PREFIX}{session}"


def _redis_admin_reset_key(token: str) -> str:
    return f"{REDIS_ADMIN_RESET_PREFIX}{token}"


def _generate_reset_code() -> str:
    """Generate a 7-character reset code like X7K-29M.
    Avoids visually ambiguous characters (0, O, I, 1) so it's easy to read."""
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    a = "".join(secrets.choice(chars) for _ in range(3))
    b = "".join(secrets.choice(chars) for _ in range(3))
    return f"{a}-{b}"


def _store_reset_code(token: str, code: str) -> None:
    expires = datetime.now(timezone.utc).timestamp() + ADMIN_RESET_TTL_SECONDS
    entry = {"code": code, "expires": expires}
    ADMIN_RESET_CODES[token] = entry
    if _use_redis_state():
        try:
            redis_client.setex(_redis_admin_reset_key(token), ADMIN_RESET_TTL_SECONDS, json.dumps(entry))
        except Exception as exc:
            logger.warning("Could not store reset code in Redis: %s", exc)


def _validate_and_consume_reset_code(token: str, code: str) -> bool:
    """Return True and consume the code if it is valid and unexpired."""
    now_ts = datetime.now(timezone.utc).timestamp()

    if _use_redis_state():
        try:
            raw = redis_client.get(_redis_admin_reset_key(token))
            if raw:
                entry = json.loads(raw)
                if entry.get("expires", 0) > now_ts and entry.get("code", "") == code:
                    redis_client.delete(_redis_admin_reset_key(token))
                    ADMIN_RESET_CODES.pop(token, None)
                    return True
            return False
        except Exception as exc:
            logger.warning("Redis reset code validation failed, falling back: %s", exc)

    entry = ADMIN_RESET_CODES.get(token)
    if entry and entry.get("expires", 0) > now_ts and entry.get("code", "") == code:
        ADMIN_RESET_CODES.pop(token, None)
        return True
    return False


def _send_telegram(message: str) -> None:
    """Send a plain-text or HTML Telegram message. Fire-and-forget; logs but never raises."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        logger.warning("Telegram not configured — skipping message")
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        data = urllib.parse.urlencode({
            "chat_id": TELEGRAM_CHAT_ID,
            "text": message,
            "parse_mode": "HTML",
        }).encode("utf-8")
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=15) as resp:
            logger.info("Telegram message sent: %s", resp.read()[:80])
    except Exception as exc:
        logger.error("Telegram delivery failed: %s", exc)


def _redis_admin_login_attempt_key(client_ip: str) -> str:
    safe_ip = re.sub(r"[^A-Za-z0-9_.:-]", "_", client_ip or "unknown")
    return f"{REDIS_ADMIN_LOGIN_ATTEMPT_PREFIX}{safe_ip}"


@contextmanager
def _redis_queue_lock():
    if redis_client is None:
        raise HTTPException(status_code=503, detail="Redis queue is unavailable")

    lock = redis_client.lock(REDIS_QUEUE_LOCK_KEY, timeout=10, blocking_timeout=5)
    acquired = False
    try:
        acquired = lock.acquire(blocking=True)
        if not acquired:
            raise HTTPException(status_code=503, detail="OTP queue is busy. Try again.")
        yield
    finally:
        if acquired:
            try:
                lock.release()
            except Exception:
                pass


def _parse_redis_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value).astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        return _utcnow_naive()


def _redis_queue_tokens() -> List[str]:
    if redis_client is None:
        return []
    return [str(token).upper() for token in redis_client.lrange(REDIS_QUEUE_KEY, 0, -1)]


def _redis_get_claim(token: str) -> Optional[Dict[str, Any]]:
    if redis_client is None:
        return None
    row = redis_client.hgetall(_redis_claim_key(token))
    if not row:
        return None
    return {
        "token": str(row.get("token") or token).upper(),
        "name": str(row.get("name") or users.get(token, {}).get("name", "")),
        "email": str(row.get("email") or users.get(token, {}).get("email", "")),
        "claimed_at": _parse_redis_datetime(str(row.get("claimed_at") or "")),
    }


def _redis_remove_claim(token: str) -> None:
    if redis_client is None:
        return
    token = token.upper()
    redis_client.lrem(REDIS_QUEUE_KEY, 0, token)
    redis_client.delete(_redis_claim_key(token))


def _redis_purge_expired_claims() -> None:
    if redis_client is None:
        return

    now = _utcnow_naive()
    while True:
        token = redis_client.lindex(REDIS_QUEUE_KEY, 0)
        if not token:
            return
        token = str(token).upper()
        claim = _redis_get_claim(token)
        if not claim:
            redis_client.lpop(REDIS_QUEUE_KEY)
            redis_client.delete(_redis_claim_key(token))
            continue

        age = (now - claim["claimed_at"]).total_seconds()
        if age <= CLAIM_EXPIRY_SEC:
            return

        redis_client.lpop(REDIS_QUEUE_KEY)
        redis_client.delete(_redis_claim_key(token))
        audit("claim_expired", token, f"No OTP arrived within {CLAIM_EXPIRY_SEC}s - evicted from slot 1", "warn")


def _redis_get_pending_otp(token: str) -> Optional[Dict[str, Any]]:
    if redis_client is None:
        return None

    token = token.upper()
    key = _redis_pending_key(token)
    row = redis_client.hgetall(key)
    if not row:
        return None

    ttl = redis_client.ttl(key)
    if ttl == -2:
        return None
    if ttl == -1:
        arrived_at = _parse_redis_datetime(str(row.get("arrived_at") or ""))
        age = (_utcnow_naive() - arrived_at).total_seconds()
        ttl = max(0, int(OTP_DISPLAY_SEC - age))
        redis_client.expire(key, ttl)

    return {
        "otp": str(row.get("otp") or "-"),
        "arrived_at": _parse_redis_datetime(str(row.get("arrived_at") or "")),
        "expires_in": max(0, int(ttl)),
    }


def _redis_set_pending_otp(token: str, otp: str) -> None:
    if redis_client is None:
        return

    key = _redis_pending_key(token.upper())
    redis_client.hset(key, mapping={"otp": otp, "arrived_at": _now_iso()})
    redis_client.expire(key, OTP_DISPLAY_SEC)


def _redis_delete_pending_otp(token: str) -> bool:
    if redis_client is None:
        return False
    return bool(redis_client.delete(_redis_pending_key(token.upper())))


def _redis_add_claim(token: str) -> Dict[str, Any]:
    if redis_client is None:
        raise HTTPException(status_code=503, detail="Redis queue is unavailable")

    token = token.upper()
    now = _utcnow_naive()
    claim = {
        "token": token,
        "name": users[token]["name"],
        "email": users[token]["email"],
        "claimed_at": now,
    }
    redis_client.hset(_redis_claim_key(token), mapping={
        "token": token,
        "name": claim["name"],
        "email": claim["email"],
        "claimed_at": now.replace(tzinfo=timezone.utc).isoformat(),
    })
    redis_client.rpush(REDIS_QUEUE_KEY, token)
    return claim


def _redis_queue_position(token: str) -> int:
    token = token.upper()
    for index, queued_token in enumerate(_redis_queue_tokens(), start=1):
        if queued_token == token:
            return index
    return 0


def _redis_pop_next_claim() -> Optional[Dict[str, Any]]:
    if redis_client is None:
        return None

    _redis_purge_expired_claims()
    while True:
        token = redis_client.lpop(REDIS_QUEUE_KEY)
        if not token:
            return None
        token = str(token).upper()
        claim = _redis_get_claim(token)
        redis_client.delete(_redis_claim_key(token))
        if claim:
            return claim


def _redis_admin_queue() -> List[Dict[str, Any]]:
    if redis_client is None:
        return []

    _redis_purge_expired_claims()
    now = _utcnow_naive()
    rows: List[Dict[str, Any]] = []
    for index, token in enumerate(_redis_queue_tokens(), start=1):
        claim = _redis_get_claim(token)
        if not claim:
            continue
        rows.append({
            "token": claim["token"],
            "name": claim["name"],
            "email": claim["email"],
            "claimed_at": claim["claimed_at"].strftime("%Y-%m-%dT%H:%M:%SZ"),
            "expires_in": max(0, int(CLAIM_EXPIRY_SEC - (now - claim["claimed_at"]).total_seconds())),
            "position": index,
        })
    return rows


def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        raw = path.read_text(encoding="utf-8").strip()
        if not raw:
            return default
        return json.loads(raw)
    except Exception as exc:
        logger.warning("Could not read %s: %s", path, exc)
        return default


def _write_json(path: Path, payload: Any) -> None:
    """Write JSON atomically so a pod restart cannot leave a half-written file."""
    _ensure_data_dir()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    tmp_path.replace(path)


def _wizard_db() -> Dict[str, dict]:
    return _read_json(WIZARD_FILE, {})


def _save_wizard_db(db: Dict[str, dict]) -> None:
    _write_json(WIZARD_FILE, db)


def _auth_db() -> Dict[str, Any]:
    return _read_json(AUTH_FILE, {})


def _save_auth_db(db: Dict[str, Any]) -> None:
    _write_json(AUTH_FILE, db)


def _config_db() -> Dict[str, Any]:
    env_tokens = os.environ.get("ADMIN_TOKENS", "")
    env_default = [t.strip().upper() for t in env_tokens.split(",") if t.strip()] or DEFAULT_ADMIN_TOKENS
    return _read_json(CONFIG_FILE, {"admin_tokens": env_default})


def _save_config_db(db: Dict[str, Any]) -> None:
    _write_json(CONFIG_FILE, db)


def _purge_admin_sessions() -> None:
    """Expire stale in-memory admin sessions.

    Redis-backed admin sessions use native key TTLs, so there is nothing to
    purge here when Redis is active.
    """
    if _use_redis_state():
        return

    now_ts = datetime.now(timezone.utc).timestamp()
    stale = [s for s, meta in ADMIN_SESSIONS.items() if now_ts - float(meta.get("ts", 0)) > ADMIN_TTL_SECONDS]
    for session in stale:
        ADMIN_SESSIONS.pop(session, None)


def _create_admin_session(who: str) -> str:
    session = secrets.token_urlsafe(24)
    now_ts = datetime.now(timezone.utc).timestamp()
    meta = {"ts": now_ts, "who": who}

    ADMIN_SESSIONS[session] = meta

    if _use_redis_state():
        try:
            redis_client.setex(_redis_admin_session_key(session), ADMIN_TTL_SECONDS, json.dumps(meta))
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin session store is unavailable") from exc
            logger.warning("Could not write admin session to Redis; using in-memory fallback: %s", exc)

    return session


def _delete_admin_session(session: str) -> None:
    ADMIN_SESSIONS.pop(session, None)

    if _use_redis_state():
        try:
            redis_client.delete(_redis_admin_session_key(session))
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin session store is unavailable") from exc
            logger.warning("Could not delete admin session from Redis: %s", exc)


def _require_admin(session: Optional[str]) -> str:
    """Validate admin session and return the admin token (who) associated with it."""
    if not session:
        raise HTTPException(status_code=401, detail="Missing admin session")

    now_ts = datetime.now(timezone.utc).timestamp()

    if _use_redis_state():
        try:
            raw = redis_client.get(_redis_admin_session_key(session))
            if not raw:
                raise HTTPException(status_code=401, detail="Invalid or expired admin session")

            try:
                meta = json.loads(raw)
            except Exception:
                meta = {"ts": float(raw), "who": "unknown"}

            # Sliding expiration: every valid admin request refreshes the session TTL.
            meta["ts"] = now_ts
            redis_client.setex(_redis_admin_session_key(session), ADMIN_TTL_SECONDS, json.dumps(meta))
            ADMIN_SESSIONS[session] = meta
            return str(meta.get("who", "unknown"))
        except HTTPException:
            raise
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin session store is unavailable") from exc
            logger.warning("Could not validate admin session in Redis; using in-memory fallback: %s", exc)

    _purge_admin_sessions()
    meta = ADMIN_SESSIONS.get(session)
    if not meta:
        raise HTTPException(status_code=401, detail="Invalid or expired admin session")
    meta["ts"] = now_ts
    return str(meta.get("who", "unknown"))


def _model_dump(model: BaseModel) -> Dict[str, Any]:
    """Support both Pydantic v1 and v2."""
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _default_login_attempt_row(now_ts: float) -> Dict[str, Any]:
    return {"count": 0, "window_start": now_ts, "locked_until": 0.0}


def _get_login_attempt_row(client_ip: str, now_ts: float) -> Dict[str, Any]:
    if _use_redis_state():
        try:
            row = redis_client.hgetall(_redis_admin_login_attempt_key(client_ip))
            if row:
                return {
                    "count": int(row.get("count", 0) or 0),
                    "window_start": float(row.get("window_start", now_ts) or now_ts),
                    "locked_until": float(row.get("locked_until", 0.0) or 0.0),
                }
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin login-attempt store is unavailable") from exc
            logger.warning("Could not read admin login attempts from Redis; using in-memory fallback: %s", exc)

    return ADMIN_LOGIN_ATTEMPTS.get(client_ip, _default_login_attempt_row(now_ts))


def _save_login_attempt_row(client_ip: str, row: Dict[str, Any]) -> None:
    ADMIN_LOGIN_ATTEMPTS[client_ip] = row

    if _use_redis_state():
        try:
            redis_client.hset(_redis_admin_login_attempt_key(client_ip), mapping={
                "count": int(row.get("count", 0)),
                "window_start": float(row.get("window_start", 0.0)),
                "locked_until": float(row.get("locked_until", 0.0)),
            })
            ttl = max(ADMIN_LOGIN_WINDOW_SECONDS, ADMIN_LOGIN_LOCKOUT_SECONDS)
            redis_client.expire(_redis_admin_login_attempt_key(client_ip), ttl)
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin login-attempt store is unavailable") from exc
            logger.warning("Could not write admin login attempts to Redis; using in-memory fallback: %s", exc)


def _delete_login_attempt_row(client_ip: str) -> None:
    ADMIN_LOGIN_ATTEMPTS.pop(client_ip, None)

    if _use_redis_state():
        try:
            redis_client.delete(_redis_admin_login_attempt_key(client_ip))
        except Exception as exc:
            if REDIS_REQUIRED:
                raise HTTPException(status_code=503, detail="Redis admin login-attempt store is unavailable") from exc
            logger.warning("Could not delete admin login attempts from Redis: %s", exc)


def _check_login_rate_limit(request: Request) -> None:
    key = _client_ip(request)
    now_ts = datetime.now(timezone.utc).timestamp()
    row = _get_login_attempt_row(key, now_ts)

    if float(row.get("locked_until", 0.0)) > now_ts:
        raise HTTPException(status_code=429, detail="Too many failed login attempts. Try again later.")

    if now_ts - float(row.get("window_start", now_ts)) > ADMIN_LOGIN_WINDOW_SECONDS:
        row = _default_login_attempt_row(now_ts)

    _save_login_attempt_row(key, row)


def _record_login_failure(request: Request) -> None:
    key = _client_ip(request)
    now_ts = datetime.now(timezone.utc).timestamp()
    row = _get_login_attempt_row(key, now_ts)

    if now_ts - float(row.get("window_start", now_ts)) > ADMIN_LOGIN_WINDOW_SECONDS:
        row = _default_login_attempt_row(now_ts)

    row["count"] = int(row.get("count", 0)) + 1
    if row["count"] >= ADMIN_LOGIN_MAX_ATTEMPTS:
        row["locked_until"] = now_ts + ADMIN_LOGIN_LOCKOUT_SECONDS
    _save_login_attempt_row(key, row)


def _record_login_success(request: Request) -> None:
    _delete_login_attempt_row(_client_ip(request))


def _hash_wizard_secret(secret: str) -> str:
    return hashlib.sha256(secret.encode("utf-8")).hexdigest()


def _valid_wizard_owner(row: Dict[str, Any], client_secret: Optional[str]) -> bool:
    stored_hash = str(row.get("client_secret_hash") or "")
    if not stored_hash or not client_secret or len(client_secret) < WIZARD_CLIENT_SECRET_MIN_LENGTH:
        return False
    return secrets.compare_digest(stored_hash, _hash_wizard_secret(client_secret))


def _bind_wizard_client(row: Dict[str, Any], client_secret: Optional[str]) -> bool:
    """Silently bind or rebind a wizard record to the current browser client.

    The user token is the durable owner of wizard progress. The client secret is
    only a background edit marker, so stale PVC/browser state must not block a
    valid token from continuing its own wizard progress.
    """
    if not client_secret or len(client_secret) < WIZARD_CLIENT_SECRET_MIN_LENGTH:
        raise HTTPException(status_code=401, detail="Wizard client secret required")

    current_hash = _hash_wizard_secret(client_secret)
    if secrets.compare_digest(str(row.get("client_secret_hash") or ""), current_hash):
        return False

    row["client_secret_hash"] = current_hash
    row["client_bound_at"] = _now_iso()
    return True


def _public_wizard_record(row: Dict[str, Any]) -> Dict[str, Any]:
    public = dict(row)
    public.pop("client_secret_hash", None)
    return public


def _default_wizard_record(token: str) -> Dict[str, Any]:
    return {
        "token": token,
        "display_name": users[token]["name"],
        "iits_username": "",
        "adm_username": "",
        "completed": [],
        "adminCompleted": [],
        "iits_pw_date": None,
        "adm_pw_date": None,
        "vpn_date": None,
        "test_env": users[token].get("test_env", ""),
        "prod_env": users[token].get("prod_env", ""),
    }


class WizardRecord(BaseModel):
    token: str
    display_name: str = ""
    iits_username: str = ""
    adm_username: str = ""
    completed: List[str] = Field(default_factory=list)
    adminCompleted: List[str] = Field(default_factory=list)
    iits_pw_date: Optional[str] = None
    adm_pw_date: Optional[str] = None
    vpn_date: Optional[str] = None
    test_env: str = ""
    prod_env: str = ""


class UserLoginPayload(BaseModel):
    token: str
    pin: Optional[str] = None          # required for admin tokens
    confirm_pin: Optional[str] = None  # required when setting a PIN for the first time
    admin_session: Optional[str] = None  # existing session token for silent page-reload restore


class CredentialPayload(BaseModel):
    token: Optional[str] = None
    credential: str
    current: Optional[str] = None


class ConfigPayload(BaseModel):
    admin_tokens: List[str]


# -- User loading --------------------------------------------------------------
def load_users_from_excel(path: str, replace_existing: bool = True) -> int:
    """
    Reads users.xlsx. Expected columns (row 1 = headers):
      token    - 2 or 3 character unique string, e.g. AH or AHM
      name     - display name
      email    - company email address
    Optional columns:
      test_env - test environment assignment shown in the admin wizard
      prod_env - production environment assignment shown in the admin wizard
    Column names are case-insensitive. Spaces, dashes, and underscores are tolerated.
    Skipped rows are written to the audit log so IT can fix them.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    raw_headers = [
        _normalize_user_import_header(cell.value)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    duplicate_headers = sorted({header for header in raw_headers if header and raw_headers.count(header) > 1})
    if duplicate_headers:
        raise ValueError(f"users.xlsx has duplicate column(s): {', '.join(duplicate_headers)}")

    required_headers = {"token", "name", "email"}
    missing_headers = sorted(required_headers - set(raw_headers))
    if missing_headers:
        raise ValueError(f"users.xlsx missing required column(s): {', '.join(missing_headers)}")

    loaded = 0
    skipped = 0
    seen_tokens: Dict[str, int] = {}
    imported_users: Dict[str, Dict[str, str]] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue

        row_dict = dict(zip(raw_headers, row))
        token = _xlsx_text(row_dict.get("token")).upper()
        name = _xlsx_text(row_dict.get("name"))
        email = _xlsx_text(row_dict.get("email"))
        test_env = _xlsx_text(row_dict.get("test_env"))
        prod_env = _xlsx_text(row_dict.get("prod_env"))

        if len(token) == 0:
            audit("import_skipped", detail=f"Row {row_num}: empty token - name={repr(name)} email={repr(email)}", status="warn")
            skipped += 1
            continue

        if not (2 <= len(token) <= 3):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token must be 2 or 3 characters, got {len(token)} ({repr(token)})", status="warn")
            skipped += 1
            continue

        if not re.match(r"^[A-Z0-9]+$", token):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token contains invalid characters ({repr(token)}) - only letters and digits allowed", status="warn")
            skipped += 1
            continue

        if not email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: missing email address for {repr(name)}", status="warn")
            skipped += 1
            continue

        if "@" not in email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: invalid email address {repr(email)}", status="warn")
            skipped += 1
            continue

        if token in seen_tokens:
            audit("import_skipped", token=token, detail=f"Row {row_num}: duplicate token - already defined at row {seen_tokens[token]}", status="warn")
            skipped += 1
            continue

        seen_tokens[token] = row_num
        imported_users[token] = {
            "token": token,
            "name": name,
            "email": email,
            "test_env": test_env,
            "prod_env": prod_env,
        }
        loaded += 1

    if replace_existing:
        users.clear()
        users.update(imported_users)

    logger.info("Loaded %s users from %s (%s rows skipped)", loaded, path, skipped)
    if skipped > 0:
        audit("import_complete", detail=f"{loaded} users loaded, {skipped} rows skipped - check import_skipped entries above", status="warn")
    else:
        audit("import_complete", detail=f"{loaded} users loaded, no issues")
    return loaded


# -- Audit log ----------------------------------------------------------------
def audit(event: str, token: Optional[str] = None, detail: str = "", status: str = "info") -> None:
    entry = {
        "ts": _utcnow_naive().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "event": event,
        "token": token or "",
        "detail": detail,
        "status": status,
    }
    try:
        audit_path = Path(AUDIT_LOG_PATH)
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        with audit_path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry) + "\n")
    except Exception as exc:
        logger.warning("Could not write audit log: %s", exc)

    level = {"info": logging.INFO, "warn": logging.WARNING, "error": logging.ERROR}.get(status, logging.INFO)
    logger.log(level, "[%s] token=%s  %s", event, token or "-", detail)

    # Bump the matching Prometheus counter, if this event has one.
    # Keeping the increment here means both the Redis and in-memory queue paths
    # are covered by the single audit() call site each path already uses.
    counter = OTP_AUDIT_COUNTERS.get(event)
    if counter is not None:
        counter.inc()


def read_audit_log(limit: int = 200) -> list:
    limit = max(1, min(int(limit or 200), 2000))
    try:
        with Path(AUDIT_LOG_PATH).open("r", encoding="utf-8") as handle:
            tail = deque(handle, maxlen=limit)
        entries = [json.loads(line) for line in tail if line.strip()]
        return list(reversed(entries))
    except FileNotFoundError:
        return []
    except Exception as exc:
        logger.warning("Could not read audit log: %s", exc)
        return []


# -- Queue and OTP state helpers ----------------------------------------------
def purge_expired() -> None:
    """Evict the front-of-queue claim if it has exceeded CLAIM_EXPIRY_SEC."""
    if _use_redis_state():
        with _redis_queue_lock():
            _redis_purge_expired_claims()
        return

    now = _utcnow_naive()
    while claim_queue:
        age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if age > CLAIM_EXPIRY_SEC:
            expired = claim_queue.popleft()
            audit("claim_expired", expired["token"], f"No OTP arrived within {CLAIM_EXPIRY_SEC}s - evicted from slot 1", "warn")
        else:
            break


def purge_stale_otps() -> None:
    """Remove delivered OTPs that have exceeded OTP_DISPLAY_SEC."""
    if _use_redis_state():
        # Redis expires pending OTPs with key TTL. No disk/log copy is kept.
        return

    now = _utcnow_naive()
    stale = [
        token for token, value in pending_otps.items()
        if (now - value["arrived_at"]).total_seconds() > OTP_DISPLAY_SEC
    ]
    for token in stale:
        del pending_otps[token]
        audit("otp_display_expired", token, f"OTP display window closed after {OTP_DISPLAY_SEC}s")


def extract_otp(text: str) -> str:
    match = re.search(r"\b\d{4,8}\b", text)
    return match.group() if match else "-"


# -- Background task -----------------------------------------------------------
async def background_purge() -> None:
    """Runs every 15 seconds to expire stale queue entries and OTP display windows."""
    while True:
        await asyncio.sleep(15)
        purge_expired()
        purge_stale_otps()


# -- Endpoints ----------------------------------------------------------------
@app.on_event("startup")
async def startup() -> None:
    _ensure_data_dir()
    _init_redis_client()
    if os.path.exists(USERS_EXCEL_PATH):
        count = load_users_from_excel(USERS_EXCEL_PATH)
        audit("server_start", detail=f"{count} users loaded")
    else:
        logger.warning("users.xlsx not found at %s", USERS_EXCEL_PATH)
        audit("server_start", detail="No users.xlsx - POST /admin/reload-users after adding it", status="warn")
    asyncio.create_task(background_purge())


@app.get("/healthz")
async def healthz():
    return {"status": "ok"}


@app.get("/readyz")
async def readyz():
    redis_status = _redis_status()

    if REDIS_REQUIRED and redis_status != "ok":
        raise HTTPException(
            status_code=503,
            detail={
                "status": "not_ready",
                "users_loaded": len(users),
                "redis": redis_status,
            },
        )

    return {
        "status": "ok",
        "users_loaded": len(users),
        "redis": redis_status,
        "redis_required": REDIS_REQUIRED,
    }


def _user_fields(user: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "token": user["token"],
        "name": user["name"],
        "email": user["email"],
        "test_env": user.get("test_env", ""),
        "prod_env": user.get("prod_env", ""),
    }


@app.post("/user/login")
async def user_login(payload: UserLoginPayload, request: Request):
    """Validate a user token.  Admin tokens additionally require a PIN.

    Flow:
    • Non-admin token          → returns user fields immediately (no PIN involved).
    • Admin token, no pin sent → returns {requires_pin, needs_setup} so the
                                  frontend can reveal the PIN field inline.
    • Admin token + valid pin  → returns user fields + admin_session.
    • Admin token + admin_session still valid (page reload) → restores silently.
    """
    token = str(payload.token or "").strip().upper()
    if token not in users:
        audit("user_login_failed", token=token, detail="Unknown token", status="warn")
        raise HTTPException(status_code=404, detail="Token not recognised. Check with IT.")

    user = users[token]
    admin_token_list = _config_db().get("admin_tokens", DEFAULT_ADMIN_TOKENS)

    if token not in admin_token_list:
        # Regular user — no PIN needed
        audit("user_login", token=token, detail="User token validated")
        return _user_fields(user)

    # ── Admin token ────────────────────────────────────────────────────────────
    hashes = _auth_hashes(_auth_db())
    has_pin = token in hashes

    # Silent restore: existing admin session passed in on page reload
    if payload.admin_session:
        try:
            who = _require_admin(payload.admin_session)
            if who == token:
                audit("user_login", token=token, detail="Admin session restored silently")
                return {**_user_fields(user), "admin_session": payload.admin_session, "is_admin": True}
        except HTTPException:
            pass  # Session expired — fall through to PIN check

    # No PIN provided yet → tell the frontend to reveal the PIN field
    if not payload.pin:
        return {**_user_fields(user), "requires_pin": True, "needs_setup": not has_pin}

    pin = (payload.pin or "").strip()

    if not has_pin:
        # First-time setup (or post-reset)
        if len(pin) < 4:
            raise HTTPException(status_code=400, detail="PIN too short — use at least 4 characters")
        confirm = (payload.confirm_pin or "").strip()
        if pin != confirm:
            raise HTTPException(status_code=400, detail="PINs do not match — try again 🙈")
        hashes[token] = bcrypt.hashpw(pin.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        _save_auth_db({"hashes": hashes, "updated_at": _now_iso()})
        audit("admin_pin_setup", token=token, detail=f"PIN set for {token}")
    else:
        _check_login_rate_limit(request)
        if not bcrypt.checkpw(pin.encode("utf-8"), hashes[token].encode("utf-8")):
            _record_login_failure(request)
            audit("admin_auth_failed", token=token, detail=f"Wrong PIN for {token}", status="warn")
            raise HTTPException(status_code=401, detail="Wrong PIN — try again 🔒")
        _record_login_success(request)

    admin_session = _create_admin_session(token)
    audit("admin_auth_login", token=token, detail=f"Admin session created for {token}")
    return {**_user_fields(user), "admin_session": admin_session, "is_admin": True}


@app.post("/claim-otp")
async def claim_otp(request: Request):
    data = await request.json()
    token = str(data.get("token", "")).strip().upper()

    if token not in users:
        audit("claim_rejected", token, "Unknown token", "error")
        raise HTTPException(status_code=404, detail="Token not recognised. Check with your IT department.")

    if _use_redis_state():
        with _redis_queue_lock():
            _redis_purge_expired_claims()

            pending = _redis_get_pending_otp(token)
            if pending:
                return {"status": "otp_ready", "expires_in": pending["expires_in"]}

            position = _redis_queue_position(token)
            if position:
                claim = _redis_get_claim(token)
                remaining = CLAIM_EXPIRY_SEC
                if claim:
                    age = (_utcnow_naive() - claim["claimed_at"]).total_seconds()
                    remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
                audit("claim_duplicate", token, f"Already at position {position}", "warn")
                return {
                    "status": "already_queued",
                    "position": position,
                    "expires_in": remaining,
                    "queue_depth": len(_redis_queue_tokens()),
                }

            tokens = _redis_queue_tokens()
            if tokens:
                front_claim = _redis_get_claim(tokens[0])
                if front_claim:
                    front_age = (_utcnow_naive() - front_claim["claimed_at"]).total_seconds()
                    if front_age < CONCURRENT_RISK_SEC:
                        audit(
                            "concurrent_risk",
                            token,
                            f"New claim while {front_claim['token']} has been active for only {int(front_age)}s",
                            "warn",
                        )

            _redis_add_claim(token)
            position = _redis_queue_position(token)
            queue_depth = len(_redis_queue_tokens())
            wait_estimate = max(0, (position - 1) * CLAIM_EXPIRY_SEC)

            audit("claim_queued", token, f"Queue position {position} of {queue_depth}")
            return {
                "status": "queued",
                "position": position,
                "name": users[token]["name"],
                "expires_in": CLAIM_EXPIRY_SEC,
                "queue_depth": queue_depth,
                "wait_estimate": wait_estimate,
            }

    purge_expired()
    purge_stale_otps()

    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age = (_utcnow_naive() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            audit("claim_duplicate", token, f"Already at position {i + 1}", "warn")
            return {
                "status": "already_queued",
                "position": i + 1,
                "expires_in": remaining,
                "queue_depth": len(claim_queue),
            }

    if token in pending_otps:
        age = (_utcnow_naive() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {"status": "otp_ready", "expires_in": remaining}

    now = _utcnow_naive()

    if claim_queue:
        front_age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if front_age < CONCURRENT_RISK_SEC:
            audit(
                "concurrent_risk",
                token,
                f"New claim while {claim_queue[0]['token']} has been active for only {int(front_age)}s",
                "warn",
            )

    claim_queue.append({
        "token": token,
        "name": users[token]["name"],
        "email": users[token]["email"],
        "claimed_at": now,
    })

    position = len(claim_queue)
    queue_depth = len(claim_queue)
    wait_estimate = max(0, (position - 1) * CLAIM_EXPIRY_SEC)

    audit("claim_queued", token, f"Queue position {position} of {queue_depth}")
    return {
        "status": "queued",
        "position": position,
        "name": users[token]["name"],
        "expires_in": CLAIM_EXPIRY_SEC,
        "queue_depth": queue_depth,
        "wait_estimate": wait_estimate,
    }


@app.get("/claim-status/{token}")
async def claim_status(token: str):
    token = token.upper()

    if _use_redis_state():
        with _redis_queue_lock():
            _redis_purge_expired_claims()

            pending = _redis_get_pending_otp(token)
            if pending:
                return {"status": "delivered", "otp": pending["otp"], "expires_in": pending["expires_in"]}

            position = _redis_queue_position(token)
            if position:
                claim = _redis_get_claim(token)
                if claim:
                    age = (_utcnow_naive() - claim["claimed_at"]).total_seconds()
                    remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
                    wait_estimate = max(0, (position - 1) * CLAIM_EXPIRY_SEC)
                    return {
                        "status": "waiting",
                        "position": position,
                        "expires_in": remaining,
                        "queue_depth": len(_redis_queue_tokens()),
                        "wait_estimate": wait_estimate,
                    }

        for entry in read_audit_log(500):
            if entry.get("token") == token:
                if entry["event"] in ("otp_delivered", "otp_display_expired"):
                    return {"status": "done"}
                if entry["event"] == "claim_expired":
                    return {"status": "idle_expired"}
                break

        return {"status": "unknown"}

    purge_expired()
    purge_stale_otps()

    if token in pending_otps:
        age = (_utcnow_naive() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {"status": "delivered", "otp": pending_otps[token]["otp"], "expires_in": remaining}

    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age = (_utcnow_naive() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            wait_estimate = max(0, i * CLAIM_EXPIRY_SEC)
            return {
                "status": "waiting",
                "position": i + 1,
                "expires_in": remaining,
                "queue_depth": len(claim_queue),
                "wait_estimate": wait_estimate,
            }

    for entry in read_audit_log(500):
        if entry.get("token") == token:
            if entry["event"] in ("otp_delivered", "otp_display_expired"):
                return {"status": "done"}
            if entry["event"] == "claim_expired":
                return {"status": "idle_expired"}
            break

    return {"status": "unknown"}


@app.delete("/claim-otp/{token}")
async def cancel_claim(token: str):
    """Discard a delivered OTP and remove/requeue claim state for retry flows."""
    token = token.upper()

    if _use_redis_state():
        with _redis_queue_lock():
            if _redis_delete_pending_otp(token):
                audit("otp_discarded", token, "User requested retry - OTP discarded from Redis")

            position = _redis_queue_position(token)
            _redis_remove_claim(token)
            if position:
                audit("claim_cancelled", token, "Removed from queue by user")

        return {"status": "ok"}

    if token in pending_otps:
        del pending_otps[token]
        audit("otp_discarded", token, "User requested retry - OTP discarded from memory")

    global claim_queue
    before = len(claim_queue)
    claim_queue = deque(claim for claim in claim_queue if claim["token"] != token)
    if len(claim_queue) < before:
        audit("claim_cancelled", token, "Removed from queue by user")

    return {"status": "ok"}


@app.post("/sms-received")
async def sms_received(request: Request):
    if request.headers.get("X-Secret-Token", "") != SMS_SECRET_TOKEN:
        audit("sms_rejected", detail="Wrong secret token", status="error")
        raise HTTPException(status_code=401)

    data = await request.json()
    sms_body = str(data.get("body", "")).strip()
    audit("sms_received", detail=f"SMS arrived ({len(sms_body)} chars)")

    if _use_redis_state():
        with _redis_queue_lock():
            recipient = _redis_pop_next_claim()

        if not recipient:
            await asyncio.sleep(4)
            with _redis_queue_lock():
                recipient = _redis_pop_next_claim()

        if not recipient:
            audit("sms_unmatched", detail="No claimant in queue - SMS discarded", status="warn")
            return {"status": "no_claimant"}

        otp = extract_otp(sms_body)
        _redis_set_pending_otp(recipient["token"], otp)

        audit("otp_delivered", recipient["token"], "OTP ready for display - queue unblocked")
        return {"status": "delivered", "recipient": recipient["name"]}

    purge_expired()
    purge_stale_otps()

    if not claim_queue:
        await asyncio.sleep(4)
        purge_expired()
        if not claim_queue:
            audit("sms_unmatched", detail="No claimant in queue - SMS discarded", status="warn")
            return {"status": "no_claimant"}

    recipient = claim_queue.popleft()
    otp = extract_otp(sms_body)

    pending_otps[recipient["token"]] = {"otp": otp, "arrived_at": _utcnow_naive()}

    audit("otp_delivered", recipient["token"], "OTP ready for display - queue unblocked")
    return {"status": "delivered", "recipient": recipient["name"]}


@app.get("/admin/log")
async def get_log(limit: int = 200, x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    entries = read_audit_log(limit)
    return {"entries": entries, "total": len(entries)}


@app.get("/admin/queue")
async def get_queue(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)

    if _use_redis_state():
        with _redis_queue_lock():
            return {"queue": _redis_admin_queue()}

    now = _utcnow_naive()
    return {
        "queue": [
            {
                "token": claim["token"],
                "name": claim["name"],
                "email": claim["email"],
                "claimed_at": claim["claimed_at"].strftime("%Y-%m-%dT%H:%M:%SZ"),
                "expires_in": max(0, int(CLAIM_EXPIRY_SEC - (now - claim["claimed_at"]).total_seconds())),
                "position": i + 1,
            }
            for i, claim in enumerate(claim_queue)
        ]
    }


@app.get("/admin/users")
async def list_users(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    return {
        "count": len(users),
        "users": [
            {
                "token": user["token"],
                "name": user["name"],
                "email": user["email"],
                "test_env": user.get("test_env", ""),
                "prod_env": user.get("prod_env", ""),
            }
            for user in users.values()
        ],
    }




@app.get("/admin/users/status")
async def users_file_status(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    path = Path(USERS_EXCEL_PATH)
    exists = path.exists()
    stat = path.stat() if exists else None
    return {
        "exists": exists,
        "path": str(path),
        "users_loaded": len(users),
        "size_bytes": stat.st_size if stat else 0,
        "updated_at": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat() if stat else None,
        "max_size_bytes": USERS_EXCEL_MAX_BYTES,
    }


@app.post("/admin/users/upload")
async def upload_users_excel(file: UploadFile = File(...), x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)

    filename = file.filename or "users.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload must be an .xlsx file")

    content = await file.read(USERS_EXCEL_MAX_BYTES + 1)
    if len(content) > USERS_EXCEL_MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"users.xlsx is too large. Maximum size is {USERS_EXCEL_MAX_BYTES} bytes")

    target = Path(USERS_EXCEL_PATH)
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_name(target.stem + ".upload.tmp.xlsx")

    try:
        # Fail fast if the file is not a valid workbook before touching the live users.xlsx.
        openpyxl.load_workbook(BytesIO(content), read_only=True).close()
        tmp_path.write_bytes(content)
        parsed_count = load_users_from_excel(str(tmp_path), replace_existing=False)
        if parsed_count <= 0:
            raise ValueError("users.xlsx did not contain any valid users")

        tmp_path.replace(target)
        count = load_users_from_excel(str(target), replace_existing=True)
        audit("users_excel_uploaded", detail=f"{filename} uploaded by admin; {count} users loaded")
        return {
            "status": "ok",
            "filename": filename,
            "users_loaded": count,
            "size_bytes": len(content),
            "path": str(target),
        }
    except HTTPException:
        raise
    except Exception as exc:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        audit("users_excel_upload_failed", detail=str(exc), status="error")
        raise HTTPException(status_code=400, detail=f"Could not import users.xlsx: {exc}")


@app.post("/admin/reload-users")
async def reload_users(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    if not os.path.exists(USERS_EXCEL_PATH):
        raise HTTPException(status_code=404, detail=f"Not found: {USERS_EXCEL_PATH}")
    count = load_users_from_excel(USERS_EXCEL_PATH, replace_existing=True)
    audit("users_reloaded", detail=f"{count} users loaded")
    return {"status": "ok", "users_loaded": count}


# -- Wizard/admin server-backed endpoints -------------------------------------
def _auth_hashes(db: Dict[str, Any]) -> Dict[str, str]:
    """Return the per-user hash map, migrating from the legacy single-hash format if needed."""
    if "hashes" in db:
        return dict(db["hashes"])
    # Legacy: single shared password_hash — migrate to a map under the sentinel key "admin"
    if db.get("password_hash"):
        return {"admin": db["password_hash"]}
    return {}


@app.get("/admin/auth/status")
async def admin_auth_status():
    db = _auth_db()
    hashes = _auth_hashes(db)
    admin_tokens = _config_db().get("admin_tokens", DEFAULT_ADMIN_TOKENS)
    configured_tokens = [t for t in admin_tokens if t in hashes]
    return {
        "configured": bool(hashes),
        "configured_tokens": configured_tokens,
        "admin_tokens": admin_tokens,
    }


@app.post("/admin/auth/setup")
async def admin_auth_setup(payload: CredentialPayload):
    who = (payload.token or "").strip().upper()
    cred = (payload.credential or "").strip()
    if len(cred) < 4:
        raise HTTPException(status_code=400, detail="Credential too short (minimum 4 characters)")

    cfg = _config_db()
    admin_tokens = cfg.get("admin_tokens", DEFAULT_ADMIN_TOKENS)

    if not who or who not in admin_tokens:
        raise HTTPException(status_code=403, detail="Token is not in the admin tokens list")

    db = _auth_db()
    hashes = _auth_hashes(db)

    if who in hashes:
        if not payload.current:
            raise HTTPException(status_code=400, detail="Current credential required to change password")
        if not bcrypt.checkpw(payload.current.encode("utf-8"), hashes[who].encode("utf-8")):
            raise HTTPException(status_code=401, detail="Current credential incorrect")

    hashes[who] = bcrypt.hashpw(cred.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    _save_auth_db({"hashes": hashes, "updated_at": _now_iso()})
    session = _create_admin_session(who)
    audit("admin_auth_setup", detail=f"Admin credential set for {who}")
    return {"status": "ok", "session": session, "who": who}


@app.post("/admin/auth/login")
async def admin_auth_login(payload: CredentialPayload, request: Request):
    _check_login_rate_limit(request)

    who = (payload.token or "").strip().upper()
    if not who:
        raise HTTPException(status_code=400, detail="Token required")

    db = _auth_db()
    hashes = _auth_hashes(db)

    # Generic error message — do not reveal whether the token exists or not
    _invalid = HTTPException(status_code=401, detail="Invalid token or credential")

    stored = hashes.get(who)
    if not stored:
        _record_login_failure(request)
        audit("admin_auth_failed", detail=f"Login attempt for unconfigured token {who}", status="warn")
        raise _invalid

    if not bcrypt.checkpw((payload.credential or "").encode("utf-8"), stored.encode("utf-8")):
        _record_login_failure(request)
        audit("admin_auth_failed", detail=f"Incorrect credential for {who}", status="warn")
        raise _invalid

    _record_login_success(request)
    session = _create_admin_session(who)
    audit("admin_auth_login", detail=f"Admin session opened for {who}")
    return {"status": "ok", "session": session, "who": who}


@app.post("/admin/auth/logout")
async def admin_auth_logout(x_admin_session: Optional[str] = Header(default=None)):
    if x_admin_session:
        # Retrieve who before deleting so we can audit it
        try:
            meta = ADMIN_SESSIONS.get(x_admin_session, {})
            who = meta.get("who", "unknown") if isinstance(meta, dict) else "unknown"
        except Exception:
            who = "unknown"
        _delete_admin_session(x_admin_session)
        audit("admin_auth_logout", detail=f"Admin session closed for {who}")
    return {"status": "ok"}


@app.post("/admin/auth/reset-request")
async def admin_pin_reset_request(payload: CredentialPayload, request: Request):
    """Send a one-time reset code to Telegram for a locked-out admin."""
    who = (payload.token or "").strip().upper()
    if not who:
        raise HTTPException(status_code=400, detail="Token required")

    admin_token_list = _config_db().get("admin_tokens", DEFAULT_ADMIN_TOKENS)
    if who not in admin_token_list:
        # Generic response — don't reveal which tokens are admin
        raise HTTPException(status_code=404, detail="Token not found")

    _check_login_rate_limit(request)

    code = _generate_reset_code()
    _store_reset_code(who, code)

    msg = (
        f"🔑 <b>Admin PIN reset requested</b>\n\n"
        f"Token <b>{who}</b> needs a new PIN.\n\n"
        f"Reset code — tap to copy 👇\n"
        f"<code>{code}</code>\n\n"
        f"⏰ Valid for <b>15 minutes</b> · single use only\n\n"
        f"🙈 Wasn't <b>{who}</b>? Someone might be sniffing around.\n"
        f"Give another admin a heads-up! 🚨"
    )
    _send_telegram(msg)
    audit("admin_pin_reset_requested", token=who, detail=f"PIN reset code sent for {who}")
    return {"status": "ok"}


@app.post("/admin/auth/reset-confirm")
async def admin_pin_reset_confirm(payload: CredentialPayload, request: Request):
    """Validate the Telegram reset code and clear the PIN so the user can set a new one."""
    who = (payload.token or "").strip().upper()
    code = (payload.credential or "").strip().upper()

    if not who or not code:
        raise HTTPException(status_code=400, detail="Token and code required")

    _check_login_rate_limit(request)

    if not _validate_and_consume_reset_code(who, code):
        _record_login_failure(request)
        audit("admin_pin_reset_failed", token=who, detail="Invalid or expired reset code", status="warn")
        raise HTTPException(status_code=401, detail="Invalid or expired reset code 🔒")

    _record_login_success(request)
    db = _auth_db()
    hashes = _auth_hashes(db)
    hashes.pop(who, None)
    _save_auth_db({"hashes": hashes, "updated_at": _now_iso()})
    audit("admin_pin_reset_confirmed", token=who, detail=f"PIN cleared for {who} — fresh setup required")
    return {"status": "ok", "needs_setup": True}


@app.post("/admin/auth/reset-peer")
async def admin_pin_reset_peer(payload: CredentialPayload, x_admin_session: Optional[str] = Header(default=None)):
    """Logged-in admin resets another admin's PIN (clears the hash)."""
    actor = _require_admin(x_admin_session)
    target = (payload.token or "").strip().upper()

    if not target:
        raise HTTPException(status_code=400, detail="Target token required")

    admin_token_list = _config_db().get("admin_tokens", DEFAULT_ADMIN_TOKENS)
    if target not in admin_token_list:
        raise HTTPException(status_code=404, detail="Token not in admin list")

    db = _auth_db()
    hashes = _auth_hashes(db)
    hashes.pop(target, None)
    _save_auth_db({"hashes": hashes, "updated_at": _now_iso()})
    audit("admin_pin_reset_peer", token=target, detail=f"PIN reset by {actor} for {target}")
    return {"status": "ok", "target": target}


@app.get("/admin/config")
async def admin_config(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    return _config_db()


@app.post("/admin/config")
async def admin_config_save(payload: ConfigPayload, x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    tokens: List[str] = []
    for token in payload.admin_tokens:
        clean = str(token or "").strip().upper()
        if clean and clean not in tokens:
            tokens.append(clean)
    _save_config_db({"admin_tokens": tokens, "updated_at": _now_iso()})
    audit("admin_config_saved", detail=f"Configured admin tokens: {', '.join(tokens) or 'none'}")
    return {"status": "ok", "admin_tokens": tokens}


@app.post("/wizard/progress")
async def wizard_progress_save(
    payload: WizardRecord,
    x_wizard_client: Optional[str] = Header(default=None),
    x_admin_session: Optional[str] = Header(default=None),
):
    token = payload.token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")

    is_admin = False
    try:
        _require_admin(x_admin_session)
        is_admin = True
    except HTTPException:
        is_admin = False

    with WIZARD_DB_LOCK:
        db = _wizard_db()
        existing = db.get(token, {})

        if not is_admin:
            _bind_wizard_client(existing, x_wizard_client)

        row = _model_dump(payload)
        row["token"] = token
        row["updated_at"] = _now_iso()
        row["client_secret_hash"] = existing.get("client_secret_hash")
        if existing.get("client_bound_at"):
            row["client_bound_at"] = existing.get("client_bound_at")
        db[token] = row
        _save_wizard_db(db)

    audit("wizard_progress_saved", token=token, detail="Wizard profile/progress updated")
    return {"status": "ok", "record": _public_wizard_record(row)}


@app.get("/wizard/progress/{token}")
async def wizard_progress_get(
    token: str,
    x_wizard_client: Optional[str] = Header(default=None),
    x_admin_session: Optional[str] = Header(default=None),
):
    token = token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")

    is_admin = False
    try:
        _require_admin(x_admin_session)
        is_admin = True
    except HTTPException:
        is_admin = False

    with WIZARD_DB_LOCK:
        db = _wizard_db()
        row = db.get(token)

        if not row:
            row = _default_wizard_record(token)
            if not is_admin:
                _bind_wizard_client(row, x_wizard_client)
                row["updated_at"] = _now_iso()
                db[token] = row
                _save_wizard_db(db)
            return _public_wizard_record(row)

        if not is_admin:
            changed = _bind_wizard_client(row, x_wizard_client)
            if changed:
                row["updated_at"] = _now_iso()
                db[token] = row
                _save_wizard_db(db)

        return _public_wizard_record(row)


@app.get("/admin/wizard")
async def admin_wizard(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    db = _wizard_db()
    merged = []
    for token, user in sorted(users.items()):
        rec = db.get(token, {})
        merged.append({
            "token": token,
            "display_name": rec.get("display_name") or user.get("name", ""),
            "email": user.get("email", ""),
            "iits_username": rec.get("iits_username", ""),
            "adm_username": rec.get("adm_username", ""),
            "completed": rec.get("completed", []),
            "adminCompleted": rec.get("adminCompleted", []),
            "iits_pw_date": rec.get("iits_pw_date"),
            "adm_pw_date": rec.get("adm_pw_date"),
            "vpn_date": rec.get("vpn_date"),
            "test_env": rec.get("test_env") or user.get("test_env", ""),
            "prod_env": rec.get("prod_env") or user.get("prod_env", ""),
            "updated_at": rec.get("updated_at"),
        })
    return {"users": merged}


@app.post("/api/onboard/notify")
async def onboard_notify(request: Request):
    payload = await request.json()
    token = str(payload.get("token", "") or "").strip().upper() or None
    detail = json.dumps(payload, sort_keys=True)[:500]
    audit("onboard_notify", token=token, detail=detail)
    return {"status": "ok", "received": payload, "ts": _now_iso()}


@app.get("/guide.html", include_in_schema=False)
def serve_guide_html():
    guide_path = FRONTEND_DIR / "guide.html"
    if not guide_path.exists():
        raise HTTPException(status_code=404, detail="guide.html not deployed")
    return FileResponse(guide_path)


# -- Prometheus gauge callbacks ------------------------------------------------
# The gauges defined near the top of this file are populated on demand via
# set_function. The callbacks reference _use_redis_state, claim_queue, and
# _redis_queue_tokens, which are all defined further up by this point.

def _current_queue_depth() -> int:
    """Returns the live queue depth, regardless of state backend."""
    if _use_redis_state():
        return len(_redis_queue_tokens())
    return len(claim_queue)


otp_queue_depth.set_function(_current_queue_depth)
otp_active_user.set_function(lambda: 1 if _current_queue_depth() > 0 else 0)


# Serve frontend - must be last
app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
